import re
from pathlib import Path

import pdfplumber
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side

from openpyxl.utils import get_column_letter


OUTPUT_FILE = Path("outputs/output.xlsx")


COLUMN_RENAME_MAP = {
    "file": "File",
    "account_number": "Account Number",
    "billing_date": "Billing Date",
    "service_address": "Service Address",
    "service_date_range": "Service Date Range",
    "current_charges": "Current Charges",
    "confidence": "Confidence",
}


# =========================================================
# REGEX
# =========================================================

ACCOUNT_RE = re.compile(
    r"(?:Account\s*Number|AccountNumber|Acct\s*(?:Number|No))\s*[:\-]?\s*([A-Z0-9\-]+)",
    re.IGNORECASE,
)

BILLING_DATE_RE = re.compile(
    r"(?:Billing\s*Date|BillingDate)\s*[:\-]?\s*([0-9]{1,2}[\/\-.][0-9]{1,2}[\/\-.][0-9]{2,4})",
    re.IGNORECASE,
)

ADDRESS_RE = re.compile(
    r"^([0-9A-Z][0-9A-Z\s,\.\-]*ST JOSEPH MO(?:\s*\d{5}(?:-\d{4})?)?)$",
    re.IGNORECASE | re.MULTILINE,
)

SERVICE_DATE_RE = re.compile(
    r"service\s*from\s*([0-9]{1,2}[\/\-.][0-9]{1,2}[\/\-.][0-9]{2,4})\s*"
    r"(?:to|through|thru|[-–—])\s*([0-9]{1,2}[\/\-.][0-9]{1,2}[\/\-.][0-9]{2,4})",
    re.IGNORECASE,
)

CHARGES_RE = re.compile(
    r"(?:Current\s*Charges|CurrentCharges|Current\s*Charge|CurrentCharge)"
    r"[^\n\r\$]{0,150}\$([0-9,]+\.\d{2})",
    re.IGNORECASE,
)


# =========================================================
# PDF EXTRACTION
# =========================================================

def extract_text(filepath: Path):

    text_lines = []

    with pdfplumber.open(filepath) as pdf:

        for page in pdf.pages:

            text = page.extract_text() or ""

            if text:
                text_lines.append(text)

    return "\n".join(text_lines)


# =========================================================
# GLOBAL FIELDS
# =========================================================

def extract_global_fields(full_text: str):

    account_match = ACCOUNT_RE.search(full_text)

    billing_date_match = BILLING_DATE_RE.search(full_text)

    account_number = (
        account_match.group(1).strip()
        if account_match else "NOT FOUND"
    )

    billing_date = (
        billing_date_match.group(1).strip()
        if billing_date_match else "NOT FOUND"
    )

    return account_number, billing_date


# =========================================================
# ADDRESS SECTIONS
# =========================================================

def find_address_sections(full_text: str):

    return list(ADDRESS_RE.finditer(full_text))


# =========================================================
# CONFIDENCE SCORING
# =========================================================

def calculate_confidence(fields):

    missing_count = 0

    for field in fields:

        if "NOT FOUND" in str(field):
            missing_count += 1

    if missing_count == 0:
        return "HIGH"

    elif missing_count == 1:
        return "MEDIUM"

    return "LOW"


# =========================================================
# PARSE SECTION
# =========================================================

def parse_section(
    section_text: str,
    service_address: str,
    fallback_billing_date: str = "NOT FOUND"
):

    billing_date_match = BILLING_DATE_RE.search(section_text)

    service_date_match = SERVICE_DATE_RE.search(section_text)

    current_charges_match = CHARGES_RE.search(section_text)

    billing_date = (
        billing_date_match.group(1).strip()
        if billing_date_match
        else fallback_billing_date
    )

    service_date_range = (
        f"{service_date_match.group(1).strip()} to "
        f"{service_date_match.group(2).strip()}"
        if service_date_match
        else "NOT FOUND"
    )

    current_charges = (
        current_charges_match.group(1)
        if current_charges_match
        else "NOT FOUND"
    )

    confidence = calculate_confidence([
        billing_date,
        service_address,
        service_date_range,
        current_charges,
    ])

    return {
        "billing_date": billing_date,
        "service_address": service_address,
        "service_date_range": service_date_range,
        "current_charges": current_charges,
        "confidence": confidence,
    }


# =========================================================
# MAIN EXTRACTION
# =========================================================

def extract_electric_bill(filepath: Path):

    full_text = extract_text(filepath)

    account_number, global_billing_date = extract_global_fields(full_text)

    address_matches = find_address_sections(full_text)

    rows = []

    for index, match in enumerate(address_matches):

        start = match.start()

        end = (
            address_matches[index + 1].start()
            if index + 1 < len(address_matches)
            else len(full_text)
        )

        section_text = full_text[start:end]

        service_address = match.group(1).strip()

        section_data = parse_section(
            section_text,
            service_address,
            fallback_billing_date=global_billing_date
        )

        rows.append({
            "file": filepath.name,
            "account_number": account_number,
            **section_data,
        })

    return rows


# =========================================================
# PRETTY EXCEL
# =========================================================

def create_pretty_excel(df: pd.DataFrame):

    clean_df = df.copy()

    clean_df["current_charges"] = pd.to_numeric(
        clean_df["current_charges"]
        .astype(str)
        .str.replace(",", "", regex=False)
        .replace({"NOT FOUND": ""}),
        errors="coerce",
    )

    pretty_df = clean_df.rename(columns=COLUMN_RENAME_MAP)

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        # =====================================================
        # DETAILS SHEET
        # =====================================================

        pretty_df.to_excel(
            writer,
            sheet_name="Bill Details",
            index=False
        )

        # =====================================================
        # SUMMARY SHEET
        # =====================================================

        summary_df = pd.DataFrame({

            "Metric": [
                "Total PDF Files",
                "Total Rows Extracted",
                "Total Charges",
                "Average Charge",
                "High Confidence Rows",
                "Medium Confidence Rows",
                "Low Confidence Rows",
            ],

            "Value": [
                clean_df["file"].nunique(),
                len(clean_df),
                clean_df["current_charges"].sum(),
                clean_df["current_charges"].mean(),
                (clean_df["confidence"] == "HIGH").sum(),
                (clean_df["confidence"] == "MEDIUM").sum(),
                (clean_df["confidence"] == "LOW").sum(),
            ]
        })

        summary_df.to_excel(
            writer,
            sheet_name="Summary",
            index=False
        )

    wb = load_workbook(OUTPUT_FILE)

    # =========================================================
    # COLORS
    # =========================================================

    dark_blue = "0F243E"
    medium_blue = "1F4E78"
    light_blue = "EAF3FA"

    green = "C6EFCE"
    yellow = "FFF2CC"
    red = "F4CCCC"

    white = "FFFFFF"

    # =========================================================
    # FILLS
    # =========================================================

    title_fill = PatternFill(
        start_color=dark_blue,
        end_color=dark_blue,
        fill_type="solid"
    )

    header_fill = PatternFill(
        start_color=medium_blue,
        end_color=medium_blue,
        fill_type="solid"
    )

    alternate_fill = PatternFill(
        start_color=light_blue,
        end_color=light_blue,
        fill_type="solid"
    )

    high_fill = PatternFill(
        start_color=green,
        end_color=green,
        fill_type="solid"
    )

    medium_fill = PatternFill(
        start_color=yellow,
        end_color=yellow,
        fill_type="solid"
    )

    low_fill = PatternFill(
        start_color=red,
        end_color=red,
        fill_type="solid"
    )

    # =========================================================
    # FONTS
    # =========================================================

    title_font = Font(
        color=white,
        bold=True,
        size=16
    )

    header_font = Font(
        color=white,
        bold=True,
        size=11
    )

    body_font = Font(
        size=10
    )

    # =========================================================
    # ALIGNMENT
    # =========================================================

    center_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    left_alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    # =========================================================
    # BORDERS
    # =========================================================

    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # =========================================================
    # FORMAT SHEETS
    # =========================================================

    for ws in wb.worksheets:

        # -----------------------------------------------------
        # TITLE ROW
        # -----------------------------------------------------

        ws.insert_rows(1)

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=ws.max_column
        )

        title_cell = ws.cell(row=1, column=1)

        if ws.title == "Bill Details":

            title_cell.value = "Billing Extraction Report"

        else:

            title_cell.value = "Extraction Summary Dashboard"

        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center_alignment

        # -----------------------------------------------------
        # HEADER ROW
        # -----------------------------------------------------

        for cell in ws[2]:

            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # -----------------------------------------------------
        # DATA ROWS
        # -----------------------------------------------------

        for row in ws.iter_rows(min_row=3):

            for cell in row:

                cell.font = body_font
                cell.border = thin_border
                cell.alignment = left_alignment

            if row[0].row % 2 == 0:

                for cell in row:
                    cell.fill = alternate_fill

        # -----------------------------------------------------
        # AUTO WIDTH
        # -----------------------------------------------------

        for column_cells in ws.columns:

            length = 0

            column = column_cells[0].column

            for cell in column_cells:

                try:

                    length = max(
                        length,
                        len(str(cell.value))
                    )

                except:
                    pass

            adjusted_width = min(length + 4, 50)

            ws.column_dimensions[
                get_column_letter(column)
            ].width = adjusted_width

        # -----------------------------------------------------
        # FREEZE PANES
        # -----------------------------------------------------

        ws.freeze_panes = "A3"

        # -----------------------------------------------------
        # FILTERS
        # -----------------------------------------------------

        ws.auto_filter.ref = ws.dimensions

        # -----------------------------------------------------
        # CONFIDENCE COLORS
        # -----------------------------------------------------

        if ws.title == "Bill Details":

            confidence_column = None

            for cell in ws[2]:

                if cell.value == "Confidence":

                    confidence_column = cell.column

                    break

            if confidence_column:

                for row in range(3, ws.max_row + 1):

                    confidence_cell = ws.cell(
                        row=row,
                        column=confidence_column
                    )

                    value = str(confidence_cell.value).upper()

                    if value == "HIGH":

                        confidence_cell.fill = high_fill

                    elif value == "MEDIUM":

                        confidence_cell.fill = medium_fill

                    elif value == "LOW":

                        confidence_cell.fill = low_fill

                    confidence_cell.alignment = center_alignment

        # -----------------------------------------------------
        # CURRENCY FORMAT
        # -----------------------------------------------------

        if ws.title == "Bill Details":

            charge_column = None

            for cell in ws[2]:

                if cell.value == "Current Charges":

                    charge_column = cell.column

                    break

            if charge_column:

                for row in range(3, ws.max_row + 1):

                    amount_cell = ws.cell(
                        row=row,
                        column=charge_column
                    )

                    amount_cell.number_format = "$#,##0.00"

        # -----------------------------------------------------
        # SUMMARY FORMATTING
        # -----------------------------------------------------

        if ws.title == "Summary":

            for row in ws.iter_rows(min_row=3):

                metric_cell = row[0]
                value_cell = row[1]

                metric_cell.font = Font(bold=True)

                if isinstance(value_cell.value, (int, float)):

                    if "Charge" in str(metric_cell.value):

                        value_cell.number_format = "$#,##0.00"

                    else:

                        value_cell.number_format = "0"

                    value_cell.alignment = center_alignment

    wb.save(OUTPUT_FILE)