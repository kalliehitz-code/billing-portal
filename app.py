from flask import Flask, render_template, request, send_file
import os
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "output"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


@app.route("/")
def home():

    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload_files():

    print("UPLOAD STARTED")

    files = request.files.getlist("files")

    if not files:

        print("NO FILES FOUND")

        return "No files uploaded"

    extracted_rows = []

    for file in files:

        print(f"PROCESSING FILE: {file.filename}")

        filepath = os.path.join(
            UPLOAD_FOLDER,
            file.filename
        )

        file.save(filepath)

        print(f"FILE SAVED: {filepath}")

        try:

            with pdfplumber.open(filepath) as pdf:

                text = ""

                for page in pdf.pages:

                    page_text = page.extract_text()

                    if page_text:

                        text += page_text + "\n"

            account_number = "UNKNOWN"
            billing_date = "UNKNOWN"
            service_address = "UNKNOWN"
            service_date_range = "UNKNOWN"
            current_charges = "0.00"

            lines = text.split("\n")

            for line in lines:

                if "Account Number" in line:

                    parts = line.split()

                    account_number = parts[-1]

                if "Billing Date" in line:

                    parts = line.split()

                    billing_date = parts[-1]

                if "Service Address" in line:

                    service_address = line.replace(
                        "Service Address",
                        ""
                    ).strip()

                if "Service Period" in line:

                    service_date_range = line.replace(
                        "Service Period",
                        ""
                    ).strip()

                if "Current Charges" in line:

                    try:

                        amount = line.split("$")[-1]

                        current_charges = amount.replace(",", "").strip()

                    except:

                        pass

            row = {

                "file": file.filename,

                "account_number": account_number,

                "billing_date": billing_date,

                "service_address": service_address,

                "service_date_range": service_date_range,

                "current_charges": current_charges,

                "confidence": "HIGH"

            }

            print(row)

            extracted_rows.append(row)

        except Exception as e:

            print("ERROR PROCESSING FILE")

            print(e)

    if not extracted_rows:

        return "No data extracted"

    df = pd.DataFrame(extracted_rows)

    df["current_charges"] = pd.to_numeric(
        df["current_charges"],
        errors="coerce"
    )

    output_path = os.path.join(
        OUTPUT_FOLDER,
        "utility_bills.xlsx"
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name="Utility Bills"
        )

    workbook = load_workbook(output_path)

    worksheet = workbook["Utility Bills"]

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in worksheet[1]:

        cell.fill = header_fill

        cell.font = header_font

    for column_cells in worksheet.columns:

        length = max(
            len(str(cell.value))
            if cell.value else 0
            for cell in column_cells
        )

        worksheet.column_dimensions[
            column_cells[0].column_letter
        ].width = length + 5

    workbook.save(output_path)

    print("EXCEL CREATED")

    total_files = len(files)

    total_rows = len(df)

    total_charges = round(
        df["current_charges"].sum(),
        2
    )

    return render_template(

        "results.html",

        total_files=total_files,

        total_rows=total_rows,

        total_charges=total_charges

    )


@app.route("/download")
def download_file():

    return send_file(

        "output/utility_bills.xlsx",

        as_attachment=True

    )


if __name__ == "__main__":

    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )